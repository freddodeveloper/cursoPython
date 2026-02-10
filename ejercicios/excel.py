from openpyxl import Workbook, load_workbook

NOMBRE_ARCHIVO = "mi_reporte.xlsx"


def crear_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append(["ID", "Producto", "Precio", "Cantidad"])

    datos = [
        [1, "Laptop", 1200, 5],
        [2, "Mouse", 25, 50],
        [3, "Teclado", 45, 30]
    ]

    for fila in datos:
        ws.append(fila)

    wb.save(NOMBRE_ARCHIVO)
    print(f"Archivo '{NOMBRE_ARCHIVO}' creado exitosamente.")


def leer_excel():
    try:
        wb = load_workbook(NOMBRE_ARCHIVO)
        ws = wb["Ventas"]

        print("\n--- Leyendo datos del Excel ---")

        for fila in ws.iter_rows(values_only=True):
            print(fila)

    except FileNotFoundError:
        print("El archivo no existe. Ejecuta primero la creación.")


if __name__ == "__main__":
    crear_excel()
    leer_excel()
